import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
from pprint import pprint

def parse_hulc_xml(xml_file):
    """Parse HULC XML file and extract relevant data"""
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    data = {
        'general_data': [],
        'materials': [],
        'layers': [],
        'windows': [],
        'spaces': [],
        'walls': [],
        'thermal_bridges': [],
        'schedules': [],
        'pumps': [],
        'circulation_loops': [],
        'chillers': [],
        'boilers': [],
        'water_heaters': [],
        'heat_rejection': [],
        'hvac_systems': [],
        'zones': [],
        'fuel_meters': [],
        'elec_meters': []
    }
    
    # Extract General Data
    general = root.find('DatosGenerales')
    if general is not None:
        gen_data = {
            'Referencia Catastral': general.findtext('refCatastral', ''),
            'Tipo Vivienda': general.findtext('tipoVivienda', ''),
            'Comunidad': general.findtext('comunidad', ''),
            'Provincia': general.findtext('provincia', ''),
            'Población': general.findtext('poblacion', ''),
            'Zona Climática': general.findtext('zonaClimatica', ''),
            'Altitud': general.findtext('altitud', ''),
            'Tipo Uso': general.findtext('tipoUso', ''),
            'Clase Higrometría': general.findtext('claseHigrometria', ''),
            'Normativa Edificación': general.findtext('normativaEdificacion', ''),
            'Normativa Instalaciones': general.findtext('normativaInstalaciones', '')
        }
        data['general_data'].append(gen_data)
    
    # Parse LIDER data from CDATA section
    entrada_grafica = root.find('EntradaGraficaLIDER')
    if entrada_grafica is not None and entrada_grafica.text:
        lider_text = entrada_grafica.text
        
        # Extract materials
        materials = extract_materials(lider_text)
        data['materials'] = materials
        
        # Extract layers (constructions)
        layers = extract_layers(lider_text)
        data['layers'] = layers
        
        # Extract windows/gaps
        windows = extract_windows(lider_text)
        data['windows'] = windows
        
        # Extract spaces
        spaces = extract_spaces(lider_text)
        data['spaces'] = spaces
        
        # Extract walls
        walls = extract_walls(lider_text)
        data['walls'] = walls
        
        # Extract thermal bridges
        thermal_bridges = extract_thermal_bridges(lider_text)
        data['thermal_bridges'] = thermal_bridges

    # Parse CALENER-GT system data (can have multiple Definicion_Sistema_GT sections)
    for sistema_gt in root.findall('Definicion_Sistema_GT'):
        sistema_calener = sistema_gt.find('Definicion_Sistema_CALENER_GT')
        if sistema_calener is not None and sistema_calener.text:
            calener_text = sistema_calener.text

            # Extract HVAC equipment
            data['pumps'].extend(extract_pumps(calener_text))
            data['circulation_loops'].extend(extract_circulation_loops(calener_text))
            data['chillers'].extend(extract_chillers(calener_text))
            data['boilers'].extend(extract_boilers(calener_text))
            data['water_heaters'].extend(extract_water_heaters(calener_text))
            data['heat_rejection'].extend(extract_heat_rejection(calener_text))
            data['hvac_systems'].extend(extract_hvac_systems(calener_text))
            data['zones'].extend(extract_zones(calener_text))
            data['fuel_meters'].extend(extract_fuel_meters(calener_text))
            data['elec_meters'].extend(extract_elec_meters(calener_text))

    return data

def extract_materials(text):
    """Extract material data from LIDER text"""
    materials = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= MATERIAL' in line:
            material_name = line.split('"')[1] if '"' in line else ''
            material_data = {'Name': material_name}

            # Extract properties
            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'THICKNESS' in prop_line and '=' in prop_line:
                    material_data['Thickness (m)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'CONDUCTIVITY' in prop_line and '=' in prop_line:
                    material_data['Conductivity (W/mK)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'DENSITY' in prop_line and '=' in prop_line:
                    material_data['Density (kg/m³)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'SPECIFIC-HEAT' in prop_line and '=' in prop_line:
                    material_data['Specific Heat (J/kgK)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'GROUP' in prop_line and '=' in prop_line:
                    material_data['Group'] = prop_line.split('"')[1] if '"' in prop_line else ''
                j += 1

            if len(material_data) > 1:
                if material_name in materials:
                    print(f"Warning: Duplicate material found and discarded: '{material_name}'")
                else:
                    materials[material_name] = material_data
            i = j
        i += 1

    return list(materials.values())

def extract_layers(text):
    """Extract layer/construction data with materials and thicknesses side by side"""
    layers = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= LAYERS' in line:
            layer_name = line.split('"')[1] if '"' in line else ''
            layer_data = {'Name': layer_name}

            materials_list = []
            thickness_list = []

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'MATERIAL' in prop_line and '=' in prop_line and '(' in prop_line:
                    # Extract material list
                    start = prop_line.find('(')
                    end = prop_line.find(')')
                    if start != -1 and end != -1:
                        materials = prop_line[start+1:end].replace('"', '')
                        materials_list = [m.strip() for m in materials.split(',')]
                elif 'THICKNESS' in prop_line and '=' in prop_line and '(' in prop_line:
                    start = prop_line.find('(')
                    end = prop_line.find(')')
                    if start != -1 and end != -1:
                        thickness = prop_line[start+1:end]
                        thickness_list = [safe_float_convert(t.strip()) for t in thickness.split(',')]
                j += 1

            # Create columns for each layer
            if materials_list and thickness_list:
                # Pair materials with thicknesses
                paired = list(zip(materials_list, thickness_list))
                for idx, (mat, thick) in enumerate(paired, 1):
                    layer_data[f'Layer {idx} Material'] = mat
                    layer_data[f'Layer {idx} Thickness (m)'] = thick

                # Calculate total thickness
                try:
                    total_thickness = sum(float(t) for t in thickness_list if isinstance(t, (int, float)) or str(t).replace('.','').replace('-','').isdigit())
                    layer_data['Total Thickness (m)'] = round(total_thickness, 3)
                except:
                    layer_data['Total Thickness (m)'] = 'N/A'

            if len(layer_data) > 1:
                if layer_name in layers:
                    print(f"Warning: Duplicate layer found and discarded: '{layer_name}'")
                else:
                    layers[layer_name] = layer_data
            i = j
        i += 1

    return list(layers.values())

def extract_windows(text):
    """Extract window/gap data"""
    windows = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= GAP' in line:
            window_name = line.split('"')[1] if '"' in line else ''
            window_data = {'Name': window_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'GLASS-TYPE' in prop_line and '=' in prop_line:
                    window_data['Glass Type'] = prop_line.split('"')[1] if '"' in prop_line else ''
                elif 'NAME-FRAME' in prop_line and '=' in prop_line:
                    window_data['Frame'] = prop_line.split('"')[1] if '"' in prop_line else ''
                elif 'TRANSMITANCIA' in prop_line and '=' in prop_line:
                    window_data['U-value (W/m²K)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'SHADING-COEF' in prop_line and '=' in prop_line:
                    window_data['Shading Coef'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'PORCENTAGE' in prop_line and '=' in prop_line:
                    window_data['Frame %'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(window_data) > 1:
                if window_name in windows:
                    print(f"Warning: Duplicate window found and discarded: '{window_name}'")
                else:
                    windows[window_name] = window_data
            i = j
        i += 1

    return list(windows.values())

def extract_spaces(text):
    """Extract space data"""
    spaces = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= SPACE' in line:
            space_name = line.split('"')[1] if '"' in line else ''
            space_data = {'Name': space_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'HEIGHT' in prop_line and '=' in prop_line and 'SPACE-HEIGHT' not in prop_line:
                    space_data['Height (m)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'TYPE' in prop_line and '=' in prop_line and 'SPACE-TYPE' not in prop_line:
                    space_data['Type'] = prop_line.split('=')[1].strip()
                elif 'SPACE-TYPE' in prop_line and '=' in prop_line:
                    space_data['Space Type'] = prop_line.split('"')[1] if '"' in prop_line else ''
                elif 'POWER' in prop_line and '=' in prop_line:
                    space_data['Power (W/m²)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(space_data) > 1:
                if space_name in spaces:
                    print(f"Warning: Duplicate space found and discarded: '{space_name}'")
                else:
                    spaces[space_name] = space_data
            i = j
        i += 1

    return list(spaces.values())

def extract_walls(text):
    """Extract wall data"""
    walls = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= EXTERIOR-WALL' in line or '= INTERIOR-WALL' in line or '= UNDERGROUND-WALL' in line or '= ROOF' in line:
            wall_name = line.split('"')[1] if '"' in line else ''
            wall_type = 'EXTERIOR' if 'EXTERIOR-WALL' in line else ('INTERIOR' if 'INTERIOR-WALL' in line else ('UNDERGROUND' if 'UNDERGROUND-WALL' in line else 'ROOF'))
            wall_data = {'Name': wall_name, 'Type': wall_type}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'CONSTRUCTION' in prop_line and '=' in prop_line:
                    wall_data['Construction'] = prop_line.split('"')[1] if '"' in prop_line else ''
                elif 'ABSORPTANCE' in prop_line and '=' in prop_line:
                    wall_data['Absorptance'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'LOCATION' in prop_line and '=' in prop_line:
                    wall_data['Location'] = prop_line.split('=')[1].strip()
                j += 1

            if len(wall_data) > 1:
                if wall_name in walls:
                    print(f"Warning: Duplicate wall found and discarded: '{wall_name}'")
                else:
                    walls[wall_name] = wall_data
            i = j
        i += 1

    return list(walls.values())

def extract_thermal_bridges(text):
    """Extract thermal bridge data"""
    bridges = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= THERMAL-BRIDGE' in line:
            bridge_name = line.split('"')[1] if '"' in line else ''
            bridge_data = {'Name': bridge_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TTL' in prop_line and '=' in prop_line:
                    bridge_data['Psi (W/mK)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'LONG-TOTAL' in prop_line and '=' in prop_line:
                    bridge_data['Length (m)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'TYPE' in prop_line and '=' in prop_line:
                    bridge_data['Type'] = prop_line.split('=')[1].strip()
                j += 1

            if len(bridge_data) > 1:
                if bridge_name in bridges:
                    print(f"Warning: Duplicate thermal bridge found and discarded: '{bridge_name}'")
                else:
                    bridges[bridge_name] = bridge_data
            i = j
        i += 1

    return list(bridges.values())

def extract_pumps(text):
    """Extract pump data from CALENER-GT text"""
    pumps = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= PUMP' in line:
            pump_name = line.split('"')[1] if '"' in line else ''
            pump_data = {'Name': pump_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'HEAD' in prop_line and '=' in prop_line and 'MECH' not in prop_line:
                    pump_data['Head (m)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'MECH-EFF' in prop_line and '=' in prop_line:
                    pump_data['Mech Efficiency'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'MOTOR-EFF' in prop_line and '=' in prop_line:
                    pump_data['Motor Efficiency'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'CAP-CTRL' in prop_line and '=' in prop_line:
                    pump_data['Control'] = prop_line.split('=')[1].strip()
                elif 'C-C-FLOW' in prop_line and '=' in prop_line:
                    pump_data['Flow (l/h)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(pump_data) > 1:
                if pump_name in pumps:
                    print(f"Warning: Duplicate pump found and discarded: '{pump_name}'")
                else:
                    pumps[pump_name] = pump_data
            i = j
        i += 1

    return list(pumps.values())

def extract_circulation_loops(text):
    """Extract circulation loop data from CALENER-GT text"""
    loops = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= CIRCULATION-LOOP' in line:
            loop_name = line.split('"')[1] if '"' in line else ''
            loop_data = {'Name': loop_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    loop_data['Type'] = prop_line.split('=')[1].strip()
                elif 'LOOP-DESIGN-DT' in prop_line and '=' in prop_line:
                    loop_data['Design DT (K)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'HEAT-SETPT-T' in prop_line and '=' in prop_line:
                    loop_data['Heat Setpoint (°C)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'LOOP-PUMP' in prop_line and '=' in prop_line:
                    loop_data['Pump'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'DHW-INLET-T' in prop_line and '=' in prop_line:
                    loop_data['DHW Inlet T (°C)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-C-PROCESS-FLOW' in prop_line and '=' in prop_line:
                    loop_data['Process Flow (l/h)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(loop_data) > 1:
                if loop_name in loops:
                    print(f"Warning: Duplicate circulation loop found and discarded: '{loop_name}'")
                else:
                    loops[loop_name] = loop_data
            i = j
        i += 1

    return list(loops.values())

def extract_chillers(text):
    """Extract chiller data from CALENER-GT text"""
    chillers = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= CHILLER' in line:
            chiller_name = line.split('"')[1] if '"' in line else ''
            chiller_data = {'Name': chiller_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line and 'CONDENSER-TYPE' not in prop_line:
                    chiller_data['Type'] = prop_line.split('=')[1].strip()
                elif 'CHW-LOOP' in prop_line and '=' in prop_line:
                    chiller_data['CHW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'CHW-DT' in prop_line and '=' in prop_line:
                    chiller_data['CHW DT (K)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'CONDENSER-TYPE' in prop_line and '=' in prop_line:
                    chiller_data['Condenser Type'] = prop_line.split('=')[1].strip()
                elif 'CW-LOOP' in prop_line and '=' in prop_line:
                    chiller_data['CW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'C-C-CAPACITY' in prop_line and '=' in prop_line:
                    chiller_data['Capacity (kW)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-NUM-OF-UNITS' in prop_line and '=' in prop_line:
                    chiller_data['Number of Units'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(chiller_data) > 1:
                if chiller_name in chillers:
                    print(f"Warning: Duplicate chiller found and discarded: '{chiller_name}'")
                else:
                    chillers[chiller_name] = chiller_data
            i = j
        i += 1

    return list(chillers.values())

def extract_boilers(text):
    """Extract boiler data from CALENER-GT text"""
    boilers = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= BOILER' in line:
            boiler_name = line.split('"')[1] if '"' in line else ''
            boiler_data = {'Name': boiler_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    boiler_data['Type'] = prop_line.split('=')[1].strip()
                elif 'HW-LOOP' in prop_line and '=' in prop_line:
                    boiler_data['HW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'C-THERM-EFF-MAX' in prop_line and '=' in prop_line:
                    boiler_data['Max Thermal Efficiency'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-C-CAPACITY' in prop_line and '=' in prop_line:
                    boiler_data['Capacity (kW)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-C-SUBTYPE' in prop_line and '=' in prop_line:
                    boiler_data['Subtype'] = prop_line.split('=')[1].strip()
                j += 1

            if len(boiler_data) > 1:
                if boiler_name in boilers:
                    print(f"Warning: Duplicate boiler found and discarded: '{boiler_name}'")
                else:
                    boilers[boiler_name] = boiler_data
            i = j
        i += 1

    return list(boilers.values())

def extract_water_heaters(text):
    """Extract domestic water heater data from CALENER-GT text"""
    heaters = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= DW-HEATER' in line:
            heater_name = line.split('"')[1] if '"' in line else ''
            heater_data = {'Name': heater_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    heater_data['Type'] = prop_line.split('=')[1].strip()
                elif 'TANK-VOLUME' in prop_line and '=' in prop_line:
                    heater_data['Tank Volume (l)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'FUEL-METER' in prop_line and '=' in prop_line:
                    heater_data['Fuel Meter'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'DHW-LOOP' in prop_line and '=' in prop_line:
                    heater_data['DHW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'C-C-CAPACITY' in prop_line and '=' in prop_line:
                    heater_data['Capacity (kW)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-CATEGORY' in prop_line and '=' in prop_line:
                    heater_data['Category'] = prop_line.split('=')[1].strip()
                j += 1

            if len(heater_data) > 1:
                if heater_name in heaters:
                    print(f"Warning: Duplicate water heater found and discarded: '{heater_name}'")
                else:
                    heaters[heater_name] = heater_data
            i = j
        i += 1

    return list(heaters.values())

def extract_heat_rejection(text):
    """Extract heat rejection equipment data from CALENER-GT text"""
    equipment = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= HEAT-REJECTION' in line:
            equip_name = line.split('"')[1] if '"' in line else ''
            equip_data = {'Name': equip_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    equip_data['Type'] = prop_line.split('=')[1].strip()
                elif 'FAN-KW/CELL' in prop_line and '=' in prop_line:
                    equip_data['Fan kW/Cell'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'DESIGN-APPROACH' in prop_line and '=' in prop_line:
                    equip_data['Design Approach (K)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'DESIGN-WETBULB' in prop_line and '=' in prop_line:
                    equip_data['Design Wetbulb (°C)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'CW-LOOP' in prop_line and '=' in prop_line:
                    equip_data['CW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'C-C-CAPACITY' in prop_line and '=' in prop_line:
                    equip_data['Capacity (kW)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(equip_data) > 1:
                if equip_name in equipment:
                    print(f"Warning: Duplicate heat rejection equipment found and discarded: '{equip_name}'")
                else:
                    equipment[equip_name] = equip_data
            i = j
        i += 1

    return list(equipment.values())

def extract_hvac_systems(text):
    """Extract HVAC system data from CALENER-GT text"""
    systems = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= SYSTEM' in line:
            system_name = line.split('"')[1] if '"' in line else ''
            system_data = {'Name': system_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    system_data['Type'] = prop_line.split('=')[1].strip()
                elif 'CHW-LOOP' in prop_line and '=' in prop_line:
                    system_data['CHW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'ZONE-HW-LOOP' in prop_line and '=' in prop_line:
                    system_data['Zone HW Loop'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'C-C-COOL-CAP' in prop_line and '=' in prop_line:
                    system_data['Cooling Capacity (kW)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-C-HEAT-CAP' in prop_line and '=' in prop_line:
                    system_data['Heating Capacity (kW)'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-C-SUPPLY-FLOW' in prop_line and '=' in prop_line:
                    system_data['Supply Flow (m³/h)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(system_data) > 1:
                if system_name in systems:
                    print(f"Warning: Duplicate HVAC system found and discarded: '{system_name}'")
                else:
                    systems[system_name] = system_data
            i = j
        i += 1

    return list(systems.values())

def extract_zones(text):
    """Extract zone data from CALENER-GT text"""
    zones = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= ZONE' in line:
            zone_name = line.split('"')[1] if '"' in line else ''
            zone_data = {'Name': zone_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    zone_data['Type'] = prop_line.split('=')[1].strip()
                elif 'SPACE' in prop_line and '=' in prop_line:
                    zone_data['Space'] = prop_line.split('"')[1] if '"' in prop_line else prop_line.split('=')[1].strip()
                elif 'C-C-OA-FLOW/PER' in prop_line and '=' in prop_line:
                    zone_data['OA Flow/Person (m³/h·p)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(zone_data) > 1:
                if zone_name in zones:
                    print(f"Warning: Duplicate zone found and discarded: '{zone_name}'")
                else:
                    zones[zone_name] = zone_data
            i = j
        i += 1

    return list(zones.values())

def extract_fuel_meters(text):
    """Extract fuel meter data from CALENER-GT text"""
    meters = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= FUEL-METER' in line:
            meter_name = line.split('"')[1] if '"' in line else ''
            meter_data = {'Name': meter_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    meter_data['Type'] = prop_line.split('=')[1].strip()
                elif 'C-C-COEF-CO2' in prop_line and '=' in prop_line:
                    meter_data['CO2 Coefficient'] = safe_float_convert(prop_line.split('=')[1].strip())
                elif 'C-C-COSTE' in prop_line and '=' in prop_line:
                    meter_data['Cost (€/kWh)'] = safe_float_convert(prop_line.split('=')[1].strip())
                j += 1

            if len(meter_data) > 1:
                if meter_name in meters:
                    print(f"Warning: Duplicate fuel meter found and discarded: '{meter_name}'")
                else:
                    meters[meter_name] = meter_data
            i = j
        i += 1

    return list(meters.values())

def extract_elec_meters(text):
    """Extract electric meter data from CALENER-GT text"""
    meters = {}
    lines = text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        if '= ELEC-METER' in line:
            meter_name = line.split('"')[1] if '"' in line else ''
            meter_data = {'Name': meter_name}

            j = i + 1
            while j < len(lines) and '..' not in lines[j]:
                prop_line = lines[j].strip()
                if 'TYPE' in prop_line and '=' in prop_line:
                    meter_data['Type'] = prop_line.split('=')[1].strip()
                j += 1

            if len(meter_data) > 1:
                if meter_name in meters:
                    print(f"Warning: Duplicate electric meter found and discarded: '{meter_name}'")
                else:
                    meters[meter_name] = meter_data
            i = j
        i += 1

    return list(meters.values())

def create_excel(data, output_file):
    """Create Excel file with formatted sheets"""
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # General Data
        if data['general_data']:
            df_general = pd.DataFrame(data['general_data'])
            df_general.to_excel(writer, sheet_name='General Data', index=False)
        
        # Materials
        if data['materials']:
            df_materials = pd.DataFrame(data['materials'])
            df_materials.to_excel(writer, sheet_name='Materials', index=False)
        
        # Layers/Constructions
        if data['layers']:
            df_layers = pd.DataFrame(data['layers'])
            df_layers.to_excel(writer, sheet_name='Constructions', index=False)
        
        # Windows
        if data['windows']:
            df_windows = pd.DataFrame(data['windows'])
            df_windows.to_excel(writer, sheet_name='Windows', index=False)
        
        # Spaces
        if data['spaces']:
            df_spaces = pd.DataFrame(data['spaces'])
            df_spaces.to_excel(writer, sheet_name='Spaces', index=False)
        
        # Walls
        if data['walls']:
            df_walls = pd.DataFrame(data['walls'])
            df_walls.to_excel(writer, sheet_name='Walls', index=False)
        
        # Thermal Bridges
        if data['thermal_bridges']:
            df_bridges = pd.DataFrame(data['thermal_bridges'])
            df_bridges.to_excel(writer, sheet_name='Thermal Bridges', index=False)

        # HVAC Equipment - Pumps
        if data['pumps']:
            df_pumps = pd.DataFrame(data['pumps'])
            df_pumps.to_excel(writer, sheet_name='Pumps', index=False)

        # Circulation Loops
        if data['circulation_loops']:
            df_loops = pd.DataFrame(data['circulation_loops'])
            df_loops.to_excel(writer, sheet_name='Circulation Loops', index=False)

        # Chillers
        if data['chillers']:
            df_chillers = pd.DataFrame(data['chillers'])
            df_chillers.to_excel(writer, sheet_name='Chillers', index=False)

        # Boilers
        if data['boilers']:
            df_boilers = pd.DataFrame(data['boilers'])
            df_boilers.to_excel(writer, sheet_name='Boilers', index=False)

        # Water Heaters
        if data['water_heaters']:
            df_heaters = pd.DataFrame(data['water_heaters'])
            df_heaters.to_excel(writer, sheet_name='Water Heaters', index=False)

        # Heat Rejection Equipment
        if data['heat_rejection']:
            df_heat_rej = pd.DataFrame(data['heat_rejection'])
            df_heat_rej.to_excel(writer, sheet_name='Heat Rejection', index=False)

        # HVAC Systems
        if data['hvac_systems']:
            df_systems = pd.DataFrame(data['hvac_systems'])
            df_systems.to_excel(writer, sheet_name='HVAC Systems', index=False)

        # Zones
        if data['zones']:
            df_zones = pd.DataFrame(data['zones'])
            df_zones.to_excel(writer, sheet_name='Zones', index=False)

        # Fuel Meters
        if data['fuel_meters']:
            df_fuel = pd.DataFrame(data['fuel_meters'])
            df_fuel.to_excel(writer, sheet_name='Fuel Meters', index=False)

        # Electric Meters
        if data['elec_meters']:
            df_elec = pd.DataFrame(data['elec_meters'])
            df_elec.to_excel(writer, sheet_name='Electric Meters', index=False)

        # Format sheets
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format header
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"Excel file created successfully: {output_file}")

def process_single_file(input_file):
    """Process a single XML file and create Excel output"""
    print(f"Parsing HULC XML file: {input_file}")
    data = parse_hulc_xml(input_file)

    output_file = f"{os.path.splitext(input_file)[0]}_results.xlsx"
    print(f"Creating Excel file: {output_file}")
    create_excel(data, output_file)

    print(f"\nData extracted from {os.path.basename(input_file)}:")
    print(f"  - General data: {len(data['general_data'])} records")
    print(f"  - Materials: {len(data['materials'])} records")
    print(f"  - Constructions: {len(data['layers'])} records")
    print(f"  - Windows: {len(data['windows'])} records")
    print(f"  - Spaces: {len(data['spaces'])} records")
    print(f"  - Walls: {len(data['walls'])} records")
    print(f"  - Thermal bridges: {len(data['thermal_bridges'])} records")
    print(f"\nHVAC Equipment (CALENER-GT):")
    print(f"  - Pumps: {len(data['pumps'])} records")
    print(f"  - Circulation loops: {len(data['circulation_loops'])} records")
    print(f"  - Chillers: {len(data['chillers'])} records")
    print(f"  - Boilers: {len(data['boilers'])} records")
    print(f"  - Water heaters: {len(data['water_heaters'])} records")
    print(f"  - Heat rejection equipment: {len(data['heat_rejection'])} records")
    print(f"  - HVAC systems: {len(data['hvac_systems'])} records")
    print(f"  - Zones: {len(data['zones'])} records")
    print(f"  - Fuel meters: {len(data['fuel_meters'])} records")
    print(f"  - Electric meters: {len(data['elec_meters'])} records")
    print(f"Excel file created: {output_file}\n")

def process_directory(input_dir):
    """Process all *.ctehexml files in a directory"""
    import glob

    # Find all .ctehexml files in the directory
    pattern = os.path.join(input_dir, "*.ctehexml")
    xml_files = glob.glob(pattern)

    if not xml_files:
        print(f"No *.ctehexml files found in directory: {input_dir}")
        return

    print(f"Found {len(xml_files)} *.ctehexml files in directory: {input_dir}")

    for xml_file in xml_files:
        try:
            process_single_file(xml_file)
        except Exception as e:
            print(f"Error processing {xml_file}: {str(e)}")
            continue

import os

def safe_float_convert(value):
    """Safely convert string to float, return original if conversion fails"""
    if isinstance(value, str):
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    return value

if __name__ == "__main__":
    import argparse
    import glob

    """Command line interface for HULC to Excel conversion."""
    parser = argparse.ArgumentParser(description="Convert HULC files to Excel format")
    parser.add_argument("-i", "--input", required=True, 
                       help="Input file path (.ctedbxml) or input directory path (processes all *.ctehexml files)")

    args = parser.parse_args()

    # Get absolute path of input
    input_path = os.path.abspath(args.input)

    # Check if input exists
    if not os.path.exists(input_path):
        print(f"Error: Input path does not exist: {input_path}")
        exit(1)

    # Determine if input is a file or directory
    if os.path.isfile(input_path):
        # Process single file
        file_ext = os.path.splitext(input_path)[1].lower()

        if file_ext == '.ctehexml':
            # Process .ctedbxml file
            process_single_file(input_path)
        else:
            print(f"Error: Unsupported file extension: {file_ext}")
            print("Supported extensions: .ctehexml, .ctedbxml")
            exit(1)

    elif os.path.isdir(input_path):
        # Process directory - find all .ctehexml files
        process_directory(input_path)

    else:
        print(f"Error: Input path is neither a file nor a directory: {input_path}")
        exit(1)
        



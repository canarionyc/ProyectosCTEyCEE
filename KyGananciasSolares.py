#%% setup
import os
import csv
import openpyxl
from openpyxl import Workbook
print(os.getcwd())

from pyCTE.utils.hulc_to_excel import safe_float_convert

#%% Read the file
input_file = r'CTEHE2019\Proyectos\Ejemplo1_2526\KyGananciasSolares.txt'
#output_file = r'CTEHE2019\Proyectos\Ejemplo1_2526\KyGananciasSolares.xlsx'
output_file = f"{os.path.splitext(input_file)[0]}_results.xlsx"
print(f"Creating Excel file: {output_file}")

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Create sheets - Summary first
ws_summary = wb.create_sheet("Resumen", 0)
ws_cerramientos = wb.create_sheet("Cerramientos Opacos")
ws_huecos = wb.create_sheet("Huecos")
ws_puentes = wb.create_sheet("Puentes Térmicos")
ws_radiacion = wb.create_sheet("Radiación Solar Julio")
ws_ganancias = wb.create_sheet("Ganancias Solares Huecos")

# Headers for Cerramientos Opacos
ws_cerramientos.append([
    "Tipo", "Nombre", "Superficie (m2)", "Transmitancia (W/m2K)",
    "Factor b (-)", "Tipo de cerramiento", "Orientación", "Nombre de Construcción"
])

# Headers for Huecos
ws_huecos.append([
    "Tipo", "Nombre", "Superficie (m2)", "Transmitancia (W/m2K)",
    "Orientación", "Porcentaje de marco (%)", "Factor solar vidrio (-)",
    "No se usa (-1)", "Factor de Sombra del hueco (-)",
    "Permeabilidad del hueco (m3/hm2 a 100 Pa)", "Nombre de Hueco"
])

# Headers for Puentes Térmicos
ws_puentes.append([
    "Tipo", "Longitud (m)", "Transmitancia térmica lineal (W/mK)", "Tipo de PPTT"
])

# Headers for Radiación Solar
ws_radiacion.append([
    "Orientación", "Radiación (kWh/m2)"
])

# Headers for Ganancias Solares
ws_ganancias.append([
    "Nombre del hueco", "Orientación (grados)", "Superficie (m2)",
    "Radiación sin obstáculos (Wh/m2)", "Radiación tras obstáculos remotos (Wh/m2)",
    "Radiación tras obstáculos fachada (Wh/m2)", "Radiación tras lamas (Wh/m2)",
    "Ganancia Solar (Wh/m2)"
])

#%% Parse the file
with open(input_file, 'r', encoding='iso-8859-1') as f:
    lines = f.readlines()

mode = "factor_perdidas"
orientaciones = ["Norte", "NE", "E", "SE", "S", "SO", "O", "NO", "Horizontal"]
rad_index = 0

for line in lines:
    line = line.strip()

    # Skip comments and empty lines
    if line.startswith('###') or line.startswith('##') or line.startswith('#') or not line:
        continue

    # Check mode changes
    if "Datos para Factor de Insolación" in line or "Datos para Factor de Insolaci" in line:
        mode = "insolacion"
        continue
    if line.startswith("Coeficiente K"):
        continue
    if line.lower() == "fin":
        break

    # Parse data
    parts = [p.strip() for p in line.split(';')]

    if mode == "factor_perdidas":
        if parts[0] == "Muro":
            # Convert numeric columns to float
            converted = [parts[0], parts[1]]  # Tipo, Nombre
            converted.append(safe_float_convert(parts[2]))  # Superficie
            converted.append(safe_float_convert(parts[3]))  # Transmitancia
            converted.append(safe_float_convert(parts[4]))  # Factor b
            converted.extend(parts[5:])  # Tipo cerramiento, Orientación, Nombre Construcción
            ws_cerramientos.append(converted)
        elif parts[0] == "Ventana":
            # Convert numeric columns to float
            converted = [parts[0], parts[1]]  # Tipo, Nombre
            converted.append(safe_float_convert(parts[2]))  # Superficie
            converted.append(safe_float_convert(parts[3]))  # Transmitancia
            converted.append(parts[4])  # Orientación
            converted.append(safe_float_convert(parts[5]))  # Porcentaje marco
            converted.append(safe_float_convert(parts[6]))  # Factor solar
            converted.append(safe_float_convert(parts[7]))  # No se usa
            converted.append(safe_float_convert(parts[8]))  # Factor sombra
            converted.append(safe_float_convert(parts[9]))  # Permeabilidad
            if len(parts) > 10:
                converted.append(parts[10])  # Nombre Hueco
            ws_huecos.append(converted)
        elif parts[0] == "PPTT":
            # Convert comma to dot for decimal separator and convert to float
            converted = [parts[0]]
            converted.append(safe_float_convert(parts[1].replace(',', '.')))  # Longitud
            converted.append(safe_float_convert(parts[2].replace(',', '.')))  # Transmitancia
            if len(parts) > 3:
                converted.extend(parts[3:])  # Tipo PPTT and rest
            ws_puentes.append(converted)

    elif mode == "insolacion":
        # Check if it's radiation data (starts with digit or quote)
        if parts[0].replace(' ', '').isdigit():
            # Radiación por orientación
            orientacion_idx = int(parts[0])
            valor = float(parts[1].strip())
            ws_radiacion.append([orientaciones[orientacion_idx], valor])
        elif parts[0].startswith('"'):
            # Ganancias solares por hueco
            nombre = parts[0].strip('"')
            data = [nombre]
            for i in range(1, len(parts)):
                try:
                    data.append(float(parts[i].strip()))
                except:
                    data.append(parts[i].strip())
            ws_ganancias.append(data)

#%% Calculate summary
# Calculate total losses/gains by conduction for Cerramientos Opacos
total_cerramientos_superficie = 0
total_cerramientos_UA = 0  # U*A (W/K)

for row in ws_cerramientos.iter_rows(min_row=2, max_row=ws_cerramientos.max_row):
    superficie = row[2].value  # Superficie (m2)
    transmitancia = row[3].value  # Transmitancia (W/m2K)
    if superficie and transmitancia:
        try:
            sup = float(superficie)
            trans = float(transmitancia)
            total_cerramientos_superficie += sup
            total_cerramientos_UA += sup * trans
        except:
            pass

# Calculate total losses/gains by conduction for Huecos
total_huecos_superficie = 0
total_huecos_UA = 0  # U*A (W/K)
total_huecos_ganancias_solares = 0  # Solar gains (Wh)

for row in ws_huecos.iter_rows(min_row=2, max_row=ws_huecos.max_row):
    superficie = row[2].value  # Superficie (m2)
    transmitancia = row[3].value  # Transmitancia (W/m2K)
    if superficie and transmitancia:
        try:
            sup = float(superficie)
            trans = float(transmitancia)
            total_huecos_superficie += sup
            total_huecos_UA += sup * trans
        except:
            pass

# Calculate total solar gains for Huecos from Ganancias Solares sheet
for row in ws_ganancias.iter_rows(min_row=2, max_row=ws_ganancias.max_row):
    ganancia = row[7].value  # Ganancia Solar (Wh/m2)
    # Get the nombre to find superficie in ws_huecos
    nombre = row[0].value
    if ganancia and nombre:
        try:
            gan = float(ganancia)
            # Find corresponding hueco to get superficie
            for hueco_row in ws_huecos.iter_rows(min_row=2, max_row=ws_huecos.max_row):
                if hueco_row[1].value == nombre:
                    sup = float(hueco_row[2].value)
                    total_huecos_ganancias_solares += gan
                    break
        except:
            pass

# Calculate total losses by thermal bridges
total_puentes_longitud = 0
total_puentes_psi_L = 0  # psi*L (W/K)

for row in ws_puentes.iter_rows(min_row=2, max_row=ws_puentes.max_row):
    longitud = row[1].value  # Longitud (m)
    transmitancia = row[2].value  # Transmitancia térmica lineal (W/mK)
    if longitud and transmitancia:
        try:
            lon = float(longitud)
            trans = float(transmitancia)
            total_puentes_longitud += lon
            total_puentes_psi_L += lon * trans
        except:
            pass

# Populate summary sheet
ws_summary.append(["RESUMEN DE PÉRDIDAS Y GANANCIAS ENERGÉTICAS"])
ws_summary.append([])
ws_summary.append(["Concepto", "Valor", "Unidad"])
ws_summary.append([])

# Cerramientos Opacos
ws_summary.append(["CERRAMIENTOS OPACOS"])
ws_summary.append(["Superficie total", round(total_cerramientos_superficie, 2), "m²"])
ws_summary.append(["Pérdidas por conducción (U·A)", round(total_cerramientos_UA, 2), "W/K"])
ws_summary.append([])

# Huecos
ws_summary.append(["HUECOS"])
ws_summary.append(["Superficie total", round(total_huecos_superficie, 2), "m²"])
ws_summary.append(["Pérdidas por conducción (U·A)", round(total_huecos_UA, 2), "W/K"])
ws_summary.append(["Ganancias solares totales (Julio)", round(total_huecos_ganancias_solares, 2), "Wh"])
ws_summary.append([])

# Puentes Térmicos
ws_summary.append(["PUENTES TÉRMICOS"])
ws_summary.append(["Longitud total", round(total_puentes_longitud, 2), "m"])
ws_summary.append(["Pérdidas por conducción (ψ·L)", round(total_puentes_psi_L, 2), "W/K"])
ws_summary.append([])

# Total losses by conduction
ws_summary.append(["TOTAL PÉRDIDAS POR CONDUCCIÓN"])
total_conduccion = total_cerramientos_UA + total_huecos_UA + total_puentes_psi_L
ws_summary.append(["Total (U·A + ψ·L)", round(total_conduccion, 2), "W/K"])
ws_summary.append([])

# Format summary sheet
from openpyxl.styles import Font, PatternFill, Alignment

# Title
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_summary['A1'].font = Font(bold=True, size=14, color='FFFFFF')

# Headers
for cell in ws_summary[3]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# Section headers
for row in [5, 9, 13, 17]:
    ws_summary.cell(row, 1).font = Font(bold=True, color='1F4E78')

# Column widths
ws_summary.column_dimensions['A'].width = 40
ws_summary.column_dimensions['B'].width = 20
ws_summary.column_dimensions['C'].width = 15

# Auto-adjust column widths for other sheets
for ws in [ws_cerramientos, ws_huecos, ws_puentes, ws_radiacion, ws_ganancias]:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

#%% Save the workbook
wb.save(output_file)
print(f"Excel file created successfully: {output_file}")
print(f"\nRESUMEN:")
print(f"Cerramientos Opacos - Superficie: {total_cerramientos_superficie:.2f} m², U·A: {total_cerramientos_UA:.2f} W/K")
print(f"Huecos - Superficie: {total_huecos_superficie:.2f} m², U·A: {total_huecos_UA:.2f} W/K, Ganancias Solares: {total_huecos_ganancias_solares:.2f} Wh")
print(f"Puentes Térmicos - Longitud: {total_puentes_longitud:.2f} m, ψ·L: {total_puentes_psi_L:.2f} W/K")
print(f"TOTAL Pérdidas por Conducción: {total_conduccion:.2f} W/K")
